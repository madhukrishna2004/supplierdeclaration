import os
import time
import pandas as pd
import tkinter as tk
from tkinter import filedialog
import mysql.connector
import subprocess
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

# Database Connection
conn = mysql.connector.connect(
    host="localhost",
    user="root",
    password="krishna@9100",
    database="krishna"
)
cursor = conn.cursor()

# Ensure 'uploads' directory exists
os.makedirs("uploads", exist_ok=True)

EXCEL_FILE = "exam_file.xlsx"

# Create an Excel File with Dropdown in Column J
def create_excel():
    wb = Workbook()
    ws = wb.active
    ws.append(["Student Name", "Class", "Section", "Roll Number", "Gender", "DOB", "Contact", "Father", "Mother", "Upload File?", "File Path"])  # Headers
    
    # Add dropdown validation in column J
    dv = DataValidation(type="list", formula1='"Yes,No"', allow_blank=True)
    ws.add_data_validation(dv)

    # Apply validation to J2:J100 (adjust the range as needed)
    for row in range(2, 101):
        dv.add(ws[f'J{row}'])

    # Save the Excel file
    wb.save(EXCEL_FILE)
    print(f"Excel file '{EXCEL_FILE}' created successfully.")

    # Open Excel file for user editing
    print("\nOpening Excel file... Please update the file and save changes before closing.")
    subprocess.run(["start", EXCEL_FILE], shell=True)  # Opens Excel file on Windows

    # Wait for user to edit and save
    input("\nPress ENTER after editing & saving the Excel file...")

# Prompt User to Upload a File
def upload_file():
    root = tk.Tk()
    root.withdraw()  # Hide the main window

    file_path = filedialog.askopenfilename(title="Select a File")
    if file_path:
        save_path = os.path.join("uploads", os.path.basename(file_path))
        os.rename(file_path, save_path)  # Move file to uploads
        return save_path
    return None

# Read Excel and Handle File Upload
def process_excel():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    df = pd.read_excel(EXCEL_FILE, dtype={"Upload File?": str})  # Ensure 'Yes' is read correctly

    for index, row in df.iterrows():
        row_num = index + 2  # Adjusting for header row
        upload_choice = str(row["Upload File?"]).strip().lower()  # Ensure correct value comparison

        print(f"Checking row {row_num}: Upload File? = {upload_choice}")  # Debugging

        if upload_choice == "yes":
            print(f"Uploading file for {row['Student Name']}...")
            file_path = upload_file()

            if file_path:
                # Store in Database
                cursor.execute("INSERT INTO exam_files (student_name, file_path) VALUES (%s, %s)", (row["Student Name"], file_path))
                conn.commit()

                # Update Excel with File Path
                ws[f"K{row_num}"] = file_path
                print(f"File uploaded for {row['Student Name']}.")

    # Save Updated Excel
    wb.save(EXCEL_FILE)
    print("Excel file updated successfully.")

# Create Database Table (Run this only once)
def create_database_table():
    cursor.execute("""
    CREATE TABLE IF NOT EXISTS exam_files (
        id INT AUTO_INCREMENT PRIMARY KEY,
        student_name VARCHAR(255),
        file_path VARCHAR(255) NOT NULL
    )
    """)
    conn.commit()
    print("Database table created successfully.")

# Run Functions
create_database_table()  # Run only once
create_excel()  # Creates and opens Excel file
process_excel()  # Reads Excel, asks for file upload, and updates it

# Close Database Connection
cursor.close()
conn.close()

